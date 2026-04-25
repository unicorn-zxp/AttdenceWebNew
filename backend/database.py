"""SQLite persistent storage with project-scoped data isolation."""

import json
import os
import sqlite3
from contextlib import contextmanager

DB_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
DB_PATH = os.path.join(DB_DIR, "attendance.db")

# How long to wait (seconds) if another writer holds the lock.
# 5s is plenty for 2-3 occasional users; raises OperationalError on true deadlock.
_BUSY_TIMEOUT = 5


@contextmanager
def get_db():
    """Context manager that yields an sqlite3 connection.

    - WAL mode: allows concurrent readers while one writer is active.
    - busy_timeout: transparently retries if the DB is momentarily locked
      by another process/request (handles 2-3 concurrent users).
    """
    os.makedirs(DB_DIR, exist_ok=True)
    conn = sqlite3.connect(DB_PATH, timeout=_BUSY_TIMEOUT)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute(f"PRAGMA busy_timeout = {_BUSY_TIMEOUT * 1000}")
    conn.execute("PRAGMA foreign_keys = ON")
    try:
        yield conn
    finally:
        conn.close()


def init_db():
    """Create tables if they don't exist. Called on app startup."""
    with get_db() as conn:
        # Migration: drop old monthly_summary that lacks project_id
        old_cols = [r[1] for r in conn.execute("PRAGMA table_info(monthly_summary)").fetchall()]
        if old_cols and "project_id" not in old_cols:
            conn.execute("DROP TABLE IF EXISTS monthly_summary")

        conn.executescript("""
            CREATE TABLE IF NOT EXISTS project (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                name        TEXT    NOT NULL,
                created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS monthly_summary (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id  INTEGER NOT NULL,
                year        INTEGER NOT NULL,
                month       INTEGER NOT NULL,
                sheet_name  TEXT    NOT NULL DEFAULT '',
                people      INTEGER NOT NULL DEFAULT 0,
                total_salary   REAL NOT NULL DEFAULT 0,
                total_workdays INTEGER NOT NULL DEFAULT 0,
                total_overtime REAL NOT NULL DEFAULT 0,
                updated_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (project_id) REFERENCES project(id),
                UNIQUE(project_id, year, month)
            );

            CREATE TABLE IF NOT EXISTS calculation_result (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id    INTEGER NOT NULL,
                year          INTEGER NOT NULL,
                month         INTEGER NOT NULL,
                sheet_name    TEXT    NOT NULL DEFAULT '',
                salary_json   TEXT    NOT NULL DEFAULT '[]',
                daily_json    TEXT    NOT NULL DEFAULT '[]',
                overview_json TEXT    NOT NULL DEFAULT '{}',
                abnormal_count INTEGER NOT NULL DEFAULT 0,
                output_att_summary TEXT,
                output_ledger      TEXT,
                output_report      TEXT,
                created_at    TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (project_id) REFERENCES project(id),
                UNIQUE(project_id, year, month)
            );

            -- Seed a default project if none exists
            INSERT OR IGNORE INTO project (id, name) VALUES (1, '默认项目');
        """)
        conn.commit()


# ─── Project CRUD ───────────────────────────────────────────

def list_projects() -> list[dict]:
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM project ORDER BY id"
        ).fetchall()
    return [dict(r) for r in rows]


def create_project(name: str) -> dict:
    with get_db() as conn:
        cur = conn.execute(
            "INSERT INTO project (name) VALUES (?)", (name,)
        )
        conn.commit()
        row = conn.execute(
            "SELECT * FROM project WHERE id = ?", (cur.lastrowid,)
        ).fetchone()
    return dict(row)


def delete_project(project_id: int) -> bool:
    """Delete a project and all its associated data."""
    with get_db() as conn:
        # Cascade delete related records
        conn.execute("DELETE FROM calculation_result WHERE project_id = ?", (project_id,))
        conn.execute("DELETE FROM monthly_summary WHERE project_id = ?", (project_id,))
        count = conn.execute("DELETE FROM project WHERE id = ?", (project_id,)).rowcount
        conn.commit()
    return count > 0


# ─── Monthly Summary ───────────────────────────────────────

def upsert_month(
    project_id: int,
    year: int,
    month: int,
    sheet_name: str,
    people: int,
    total_salary: float,
    total_workdays: int,
    total_overtime: float,
):
    """Insert or update a single month's summary scoped to a project."""
    with get_db() as conn:
        conn.execute(
            """
            INSERT INTO monthly_summary
                (project_id, year, month, sheet_name, people, total_salary,
                 total_workdays, total_overtime, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(project_id, year, month) DO UPDATE SET
                sheet_name    = excluded.sheet_name,
                people        = excluded.people,
                total_salary  = excluded.total_salary,
                total_workdays = excluded.total_workdays,
                total_overtime = excluded.total_overtime,
                updated_at    = CURRENT_TIMESTAMP
            """,
            (project_id, year, month, sheet_name, people,
             total_salary, total_workdays, total_overtime),
        )
        conn.commit()


def get_annual(project_id: int, year: int) -> list[dict]:
    """Return all monthly summaries for a project-year, sorted by month."""
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM monthly_summary WHERE project_id = ? AND year = ? ORDER BY month",
            (project_id, year),
        ).fetchall()
    return [dict(r) for r in rows]


# ─── Full Calculation Results ──────────────────────────────

def save_calculation(
    project_id: int,
    year: int,
    month: int,
    sheet_name: str,
    salary_records: list[dict],
    daily_records: list[dict],
    overview: dict,
    abnormal_count: int,
    output_paths: dict[str, str] | None = None,
):
    """Save full calculation results (salary + daily JSON) to DB."""
    output_paths = output_paths or {}
    with get_db() as conn:
        conn.execute(
            """
            INSERT INTO calculation_result
                (project_id, year, month, sheet_name,
                 salary_json, daily_json, overview_json, abnormal_count,
                 output_att_summary, output_ledger, output_report)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(project_id, year, month) DO UPDATE SET
                sheet_name        = excluded.sheet_name,
                salary_json       = excluded.salary_json,
                daily_json        = excluded.daily_json,
                overview_json     = excluded.overview_json,
                abnormal_count    = excluded.abnormal_count,
                output_att_summary = excluded.output_att_summary,
                output_ledger     = excluded.output_ledger,
                output_report     = excluded.output_report,
                created_at        = CURRENT_TIMESTAMP
            """,
            (
                project_id, year, month, sheet_name,
                json.dumps(salary_records, ensure_ascii=False),
                json.dumps(daily_records, ensure_ascii=False),
                json.dumps(overview, ensure_ascii=False),
                abnormal_count,
                output_paths.get("att_summary"),
                output_paths.get("ledger"),
                output_paths.get("report"),
            ),
        )
        conn.commit()


def load_calculation(project_id: int, year: int, month: int) -> dict | None:
    """Load full calculation results from DB. Returns None if not found."""
    with get_db() as conn:
        row = conn.execute(
            "SELECT * FROM calculation_result WHERE project_id = ? AND year = ? AND month = ?",
            (project_id, year, month),
        ).fetchone()
    if not row:
        return None
    d = dict(row)
    d["salary_records"] = json.loads(d["salary_json"])
    d["daily_records"] = json.loads(d["daily_json"])
    d["overview"] = json.loads(d["overview_json"])
    return d


def list_calculations(project_id: int) -> list[dict]:
    """List all calculation results for a project."""
    with get_db() as conn:
        rows = conn.execute(
            """SELECT id, project_id, year, month, sheet_name, abnormal_count,
                      overview_json, created_at
               FROM calculation_result WHERE project_id = ? ORDER BY year, month""",
            (project_id,),
        ).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        d["overview"] = json.loads(d["overview_json"])
        del d["overview_json"]
        result.append(d)
    return result


# ─── Seed from Ledger ──────────────────────────────────────

def seed_from_ledger(project_id: int, ledger_path: str):
    """
    Read all *工资表 sheets from an uploaded ledger file and upsert their
    summaries into the database.  Only reads the ORIGINAL file (preserves
    Excel formula cache).

    Returns the number of months imported.
    """
    import re

    from openpyxl import load_workbook

    wb = load_workbook(ledger_path, data_only=True)
    count = 0

    for sheet_name in wb.sheetnames:
        if "工资表" not in sheet_name:
            continue

        m = re.search(r"(\d+)月", sheet_name)
        month_num = int(m.group(1)) if m else 0
        if month_num == 0:
            continue

        ws = wb[sheet_name]
        year_num = 2026
        date_val = ws.cell(row=3, column=18).value
        if date_val and isinstance(date_val, str):
            ym = re.search(r"(\d{4})", date_val)
            if ym:
                year_num = int(ym.group(1))

        people = 0
        total_salary = 0.0
        total_workdays = 0
        total_overtime = 0.0

        # Always sum individual data rows (rows with a name in col3).
        # Stop at 合计/总计 row or first empty-name row after data.
        for row_idx in range(6, ws.max_row + 1):
            first_val = ws.cell(row=row_idx, column=1).value
            val_str = str(first_val).strip() if first_val else ""
            if val_str in ("合计", "总计"):
                break

            name = ws.cell(row=row_idx, column=3).value
            if not name or str(name).strip() == "":
                continue

            people += 1
            salary = ws.cell(row=row_idx, column=10).value
            workdays = ws.cell(row=row_idx, column=5).value
            overtime = ws.cell(row=row_idx, column=7).value

            if salary and isinstance(salary, (int, float)):
                total_salary += float(salary)
            if workdays and isinstance(workdays, (int, float)):
                total_workdays += int(workdays)
            if overtime and isinstance(overtime, (int, float)):
                total_overtime += float(overtime)

        upsert_month(
            project_id=project_id,
            year=year_num,
            month=month_num,
            sheet_name=sheet_name,
            people=people,
            total_salary=round(total_salary, 2),
            total_workdays=int(total_workdays),
            total_overtime=round(total_overtime, 1),
        )
        count += 1

    wb.close()
    return count
