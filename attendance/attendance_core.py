"""
工地考勤工资自动计算系统 - 核心模块
适配创新智成-西安东站项目，花名册含个人工资标准
"""

import pandas as pd
import numpy as np
from datetime import datetime, time, timedelta
from typing import Dict, List, Tuple, Optional
import re
import warnings

warnings.filterwarnings('ignore')


# ============================================================
# 配置
# ============================================================

# 排除的工种（不参与工资计算，仅保留考勤）
EXCLUDED_JOB_TYPES = {'管理', '安全员', '资料员', '技术员', '安全', '资料', '材料', '分包老板'}

# 默认配置
DEFAULT_LATE_TOLERANCE = 10  # 晚班弹性补齐容差(分钟)
DEFAULT_OVERTIME_CUTOVER = time(16, 30)  # 加班分界时间


# ============================================================
# 1. 花名册解析器（新格式：班组花名册，含个人工资标准）
# ============================================================

def parse_xdz_roster(file_path: str) -> Dict[str, dict]:
    """
    解析创新智成-西安东站花名册

    读取 "创新智成 (班组花名册)  " sheet，
    提取姓名、工种、工日工资、工时工资等个人工资标准。

    Returns:
        Dict: {姓名: {工种, 工日工资, 工时工资, 备注, 性别, 身份证号码, 电话号码, 身份证地址, 合同编号}}
    """
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)

    # 查找班组花名册 sheet（处理末尾空格）
    sheet = None
    for s in wb.sheetnames:
        if '班组花名册' in s:
            sheet = wb[s]
            break

    if sheet is None:
        raise ValueError(f"未找到包含'班组花名册'的sheet，现有sheet: {wb.sheetnames}")

    roster = {}
    for row_idx in range(3, sheet.max_row + 1):
        name = sheet.cell(row=row_idx, column=3).value  # Col C: 姓名
        if not name or not str(name).strip():
            continue

        name = str(name).strip()
        job_type = sheet.cell(row=row_idx, column=6).value  # Col F: 工种
        job_type = str(job_type).strip() if job_type else ''

        daily_wage = sheet.cell(row=row_idx, column=9).value  # Col I: 工日工资
        hourly_wage = sheet.cell(row=row_idx, column=10).value  # Col J: 工时工资

        # 上报花名册需要的字段
        id_card = sheet.cell(row=row_idx, column=4).value  # Col D: 身份证号码
        gender = sheet.cell(row=row_idx, column=5).value   # Col E: 性别
        phone = sheet.cell(row=row_idx, column=7).value    # Col G: 电话号码
        address = sheet.cell(row=row_idx, column=8).value  # Col H: 身份证地址
        contract = sheet.cell(row=row_idx, column=16).value  # Col P: 合同编号

        remark = ''

        # 处理工日工资：可能为非数字（如"15000元/月"）
        daily_wage_num = None
        if daily_wage is not None:
            if isinstance(daily_wage, (int, float)):
                daily_wage_num = float(daily_wage)
            else:
                daily_wage_str = str(daily_wage).strip()
                # 尝试提取数字
                num_match = re.search(r'[\d.]+', daily_wage_str)
                if num_match and '元/月' in daily_wage_str:
                    remark = f'月薪制({daily_wage_str})，需人工核算'
                elif num_match:
                    daily_wage_num = float(num_match.group())
                else:
                    remark = f'工资标准异常({daily_wage_str})'

        # 处理工时工资
        hourly_wage_num = None
        if hourly_wage is not None:
            if isinstance(hourly_wage, (int, float)):
                hourly_wage_num = float(hourly_wage)
            else:
                num_match = re.search(r'[\d.]+', str(hourly_wage))
                if num_match:
                    hourly_wage_num = float(num_match.group())

        roster[name] = {
            '工种': job_type,
            '工日工资': daily_wage_num,
            '工时工资': hourly_wage_num,
            '备注': remark,
            '性别': str(gender).strip() if gender else '',
            '身份证号码': str(id_card).strip() if id_card else '',
            '电话号码': str(int(phone)) if isinstance(phone, (int, float)) and phone else (str(phone).strip() if phone else ''),
            '身份证地址': str(address).strip() if address else '',
            '合同编号': str(contract).strip() if contract else '',
        }

    wb.close()
    return roster


# ============================================================
# 2. 刷卡记录解析器（保持原有逻辑）
# ============================================================

def parse_punch_time(time_str: str) -> Optional[time]:
    """解析时间字符串"""
    if pd.isna(time_str):
        return None
    time_str = str(time_str).strip()
    if not time_str:
        return None

    match = re.match(r'(\d{1,2}):(\d{2})', time_str)
    if match:
        return time(int(match.group(1)), int(match.group(2)))
    return None


def parse_attendance(file_paths: List[str]) -> pd.DataFrame:
    """
    解析员工刷卡记录表

    Returns:
        DataFrame: 列包含 [姓名, 工号, 部门, 日期, 打卡时间]
    """
    all_records = []

    for file_path in file_paths:
        df = pd.read_excel(file_path, header=None)

        current_employee = None
        date_header = None
        current_row = 0

        while current_row < len(df):
            row = df.iloc[current_row]
            row_str = ' '.join([str(v) for v in row.values if pd.notna(v)])

            if '工号' in row_str and '姓名' in row_str:
                match = re.search(r'工号[：:]\s*(\d+)', row_str)
                emp_id = match.group(1) if match else None

                match = re.search(r'姓名[：:]\s*([^\s]+)', row_str)
                emp_name = match.group(1) if match else None

                match = re.search(r'部门[：:]\s*([^\s]+)', row_str)
                dept = match.group(1) if match else None

                current_employee = (emp_id, emp_name, dept)

                current_row += 1
                if current_row < len(df):
                    date_row = df.iloc[current_row]
                    date_header = []
                    for val in date_row.values:
                        try:
                            date_int = int(float(val))
                            if 1 <= date_int <= 31:
                                date_header.append(date_int)
                        except (ValueError, TypeError):
                            date_header.append(None)

                current_row += 1
                continue

            if current_employee and date_header:
                emp_id, emp_name, dept = current_employee

                punch_data = {}
                for col_idx, date_val in enumerate(date_header):
                    if date_val is None:
                        continue

                    cell_value = row.iloc[col_idx]
                    if pd.isna(cell_value):
                        continue

                    times = str(cell_value).split('\n')
                    for t in times:
                        parsed_time = parse_punch_time(t)
                        if parsed_time:
                            if date_val not in punch_data:
                                punch_data[date_val] = []
                            punch_data[date_val].append(parsed_time)

                for date_val, times in punch_data.items():
                    for punch_time in times:
                        all_records.append({
                            '姓名': emp_name,
                            '工号': emp_id,
                            '部门': dept,
                            '日期': date_val,
                            '打卡时间': punch_time
                        })

            current_row += 1

    if not all_records:
        return pd.DataFrame(columns=['姓名', '日期', '打卡时间'])

    df_result = pd.DataFrame(all_records)
    df_result = df_result.drop_duplicates(subset=['姓名', '日期', '打卡时间'])

    return df_result


# ============================================================
# 3. 工时计算引擎（保持原有逻辑）
# ============================================================

def apply_early_rounding(punch_time: time) -> time:
    """早班弹性进位: 任何 ≤ 07:40 的打卡时间都按 07:30 计"""
    minutes = punch_time.hour * 60 + punch_time.minute
    if minutes <= 7 * 60 + 40:
        return time(7, 30)
    return punch_time


def apply_late_rounding(punch_time: time, tolerance: int = DEFAULT_LATE_TOLERANCE) -> time:
    """晚班弹性补齐: 距整点/半点≤容差分钟则补齐"""
    minute = punch_time.minute

    if 30 - tolerance <= minute <= 29:
        return time(punch_time.hour, 30)
    elif 60 - tolerance <= minute <= 59:
        return time((punch_time.hour + 1) % 24, 0)

    return punch_time


def calculate_base_overtime_hours(
    times: List[time],
    overtime_cutoff: time = DEFAULT_OVERTIME_CUTOVER,
    for_settlement: bool = False
) -> Tuple[float, float, bool]:
    """
    计算每日基本工时和加班工时(以16:30为分界点)

    Returns:
        (基本工时, 加班工时, 是否异常)
    """
    if len(times) < 2:
        return 0.0, 0.0, True

    earliest = min(times)
    latest = max(times)

    if for_settlement:
        earliest = apply_early_rounding(earliest)
        latest = apply_late_rounding(latest)

    earliest_minutes = earliest.hour * 60 + earliest.minute
    latest_minutes = latest.hour * 60 + latest.minute
    cutoff_minutes = overtime_cutoff.hour * 60 + overtime_cutoff.minute

    if latest_minutes < earliest_minutes:
        latest_minutes += 24 * 60

    # 计算基本工时
    base_minutes = 0
    if earliest_minutes < cutoff_minutes:
        base_minutes = min(cutoff_minutes, latest_minutes) - earliest_minutes
        lunch_start = 12 * 60
        lunch_end = 13 * 60
        work_end = min(cutoff_minutes, latest_minutes)
        overlap_start = max(earliest_minutes, lunch_start)
        overlap_end = min(work_end, lunch_end)
        if overlap_end > overlap_start:
            base_minutes -= (overlap_end - overlap_start)

    # 计算加班工时
    overtime_minutes = 0
    if latest_minutes > cutoff_minutes:
        overtime_minutes = latest_minutes - max(cutoff_minutes, earliest_minutes)

    if base_minutes < 0:
        base_minutes = 0
    if overtime_minutes < 0:
        overtime_minutes = 0

    base_hours = base_minutes / 60.0
    overtime_hours = overtime_minutes / 60.0

    if for_settlement:
        base_hours = int(base_hours * 2) / 2
        overtime_hours = int(overtime_hours * 2) / 2

    return base_hours, overtime_hours, False


# ============================================================
# 4. 数据处理 - 按个人工资标准匹配计算
# ============================================================

def process_xdz_data(
    attendance_df: pd.DataFrame,
    roster_dict: Dict[str, dict]
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    处理考勤数据，按花名册中的个人工资标准计算工资

    异常处理：
    - 考勤有但花名册无 → 保留考勤，不计算工资，备注"花名册中无此人"
    - 花名册有但无工资标准 → 保留考勤，不计算工资，备注具体原因
    - 排除工种 → 保留考勤，不计算工资

    Returns:
        (工资汇总DataFrame, 每日考勤明细DataFrame)
    """
    grouped = attendance_df.groupby('姓名')

    salary_results = []
    daily_details = []

    for name, group in grouped:
        # 从花名册获取信息
        roster_info = roster_dict.get(name)
        job_type = roster_info['工种'] if roster_info else '未知'
        is_excluded = job_type in EXCLUDED_JOB_TYPES if roster_info else True
        has_wage = (
            roster_info
            and roster_info['工日工资'] is not None
            and roster_info['工日工资'] > 0
            and roster_info['工时工资'] is not None
            and not roster_info.get('备注', '')
        )

        daily_wage = roster_info['工日工资'] if roster_info and roster_info['工日工资'] else 0
        hourly_wage = roster_info['工时工资'] if roster_info and roster_info['工时工资'] else 0

        # 确定备注原因
        wage_note = ''
        if not roster_info:
            wage_note = '花名册中无此人'
        elif is_excluded:
            wage_note = f'排除工种({job_type})'
        elif roster_info.get('备注', ''):
            wage_note = roster_info['备注']
        elif daily_wage == 0:
            wage_note = '工日工资为0'

        by_date = group.groupby('日期')
        valid_days = 0
        total_base_hours = 0.0
        total_overtime_hours = 0.0

        for date_val, times_df in by_date:
            times = times_df['打卡时间'].tolist()
            original_times = times.copy()

            base_hours, overtime_hours, is_abnormal = calculate_base_overtime_hours(
                times, for_settlement=True
            )

            if is_abnormal:
                daily_details.append({
                    '日期': date_val,
                    '姓名': name,
                    '工种': job_type,
                    '上班打卡时间': min(original_times) if original_times else None,
                    '下班打卡时间': max(original_times) if original_times else None,
                    '当日工时': 0,
                    '基本工时': 0,
                    '加班工时': 0,
                    '当日基本工资': 0,
                    '当日加班工资': 0,
                    '当日总工资': 0,
                    '备注': '异常:单日仅一次打卡' if len(original_times) == 1 else '异常:无有效打卡'
                })
                continue

            settle_hours = base_hours + overtime_hours
            if settle_hours <= 0:
                continue

            valid_days += 1
            total_base_hours += base_hours
            total_overtime_hours += overtime_hours

            # 计算当日工资（有工资标准才计算）
            if has_wage:
                daily_base_salary = (base_hours / 8) * daily_wage
                daily_ot_salary = overtime_hours * hourly_wage
                daily_total = daily_base_salary + daily_ot_salary
            else:
                daily_base_salary = 0
                daily_ot_salary = 0
                daily_total = 0

            note = ''
            if len(original_times) > 2:
                times_str = ', '.join([t.strftime('%H:%M') for t in sorted(original_times)])
                note = f'包含多次打卡:{times_str}'

            daily_details.append({
                '日期': date_val,
                '姓名': name,
                '工种': job_type,
                '上班打卡时间': min(original_times),
                '下班打卡时间': max(original_times),
                '当日工时': settle_hours,
                '基本工时': base_hours,
                '加班工时': overtime_hours,
                '当日基本工资': round(daily_base_salary, 2),
                '当日加班工资': round(daily_ot_salary, 2),
                '当日总工资': round(daily_total, 2),
                '备注': note
            })

        if valid_days > 0:
            if has_wage:
                base_salary = (total_base_hours / 8) * daily_wage
                overtime_salary = total_overtime_hours * hourly_wage
                total_salary = base_salary + overtime_salary
            else:
                base_salary = 0
                overtime_salary = 0
                total_salary = 0

            salary_results.append({
                '序号': 0,  # 后续重新编号
                '姓名': name,
                '工种': job_type,
                '出勤工日': valid_days,
                '日工资': daily_wage,
                '加班工时': total_overtime_hours,
                '加班工资': hourly_wage,
                '工资总额': round(total_salary, 2),
                '未支付数': round(total_salary, 2),
                '备注': wage_note,
            })

    salary_df = pd.DataFrame(salary_results)

    # 重新编号
    if not salary_df.empty:
        salary_df['序号'] = range(1, len(salary_df) + 1)

    daily_df = pd.DataFrame(daily_details)

    return salary_df, daily_df


# ============================================================
# 5. 日期范围工具
# ============================================================

def get_attendance_date_range(file_paths: list):
    """
    从员工刷卡记录表中读取考勤日期范围

    Returns:
        (start_date, end_date): ((年, 月, 日), (年, 月, 日)) 或 (None, None)
    """
    all_start_dates = []
    all_end_dates = []

    for file_path in file_paths:
        try:
            df = pd.read_excel(file_path, header=None)
            for _, row in df.iterrows():
                row_str = ' '.join([str(v) for v in row if pd.notna(v)])
                if '考勤日期' in row_str:
                    match = re.search(
                        r'(\d{4})-(\d{1,2})-(\d{1,2})[^0-9]*～[^0-9]*(\d{4})-(\d{1,2})-(\d{1,2})',
                        row_str
                    )
                    if match:
                        sy, sm, sd, ey, em, ed = match.groups()
                        all_start_dates.append((int(sy), int(sm), int(sd)))
                        all_end_dates.append((int(ey), int(em), int(ed)))
                        break
        except Exception:
            continue

    if not all_start_dates:
        return None, None

    from datetime import date
    start_date = min(all_start_dates, key=lambda x: date(x[0], x[1], x[2]))
    end_date = max(all_end_dates, key=lambda x: date(x[0], x[1], x[2]))

    return start_date, end_date


def format_date_range_sheet_name(start_date: tuple, end_date: tuple) -> str:
    """生成sheet名称，如 '3月11日-4月10日工资表'"""
    sy, sm, sd = start_date
    ey, em, ed = end_date
    return f'{sm}月{sd}日-{em}月{ed}日工资表'


# ============================================================
# 6. 考勤记录汇总输出
# ============================================================

def generate_attendance_summary(daily_df: pd.DataFrame, output_path: str):
    """
    生成考勤记录汇总Excel（仅一个Sheet：每日考勤明细）

    Args:
        daily_df: 每日考勤明细DataFrame
        output_path: 输出文件路径
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = '每日考勤明细'

    # 样式
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 表头
    headers = ['日期', '姓名', '工种', '上班打卡时间', '下班打卡时间',
               '当日工时', '基本工时', '加班工时',
               '当日基本工资', '当日加班工资', '当日总工资', '备注']
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 数据
    daily_sorted = daily_df.sort_values(['姓名', '日期'])
    for _, row in daily_sorted.iterrows():
        punch_in = row['上班打卡时间'].strftime('%H:%M') if pd.notna(row['上班打卡时间']) else ''
        punch_out = row['下班打卡时间'].strftime('%H:%M') if pd.notna(row['下班打卡时间']) else ''
        ws.append([
            row['日期'], row['姓名'], row['工种'],
            punch_in, punch_out,
            row['当日工时'], row['基本工时'], row['加班工时'],
            row['当日基本工资'], row['当日加班工资'], row['当日总工资'],
            row['备注'] if pd.notna(row.get('备注')) else ''
        ])

    # 列宽
    widths = [8, 10, 8, 12, 12, 10, 10, 10, 12, 12, 12, 30]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = w

    # 数据行边框
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border

    wb.save(output_path)
    return output_path


# ============================================================
# 7. 工资台账 Sheet 生成（在现有台账文件中新增Sheet）
# ============================================================

def generate_ledger_sheet(
    ledger_path: str,
    output_path: str,
    salary_df: pd.DataFrame,
    sheet_name: str,
    start_date: tuple,
    end_date: tuple,
):
    """
    在现有工资台账文件中新增一个工资发放表Sheet

    复制模板 "2月11日-3月10日工资表" 的格式（样式/列宽/合并单元格），
    填入新数据。

    Args:
        ledger_path: 原始台账文件路径
        output_path: 输出文件路径
        salary_df: 工资汇总数据
        sheet_name: 新Sheet名称，如 "3月11日-4月10日工资表"
        start_date: (年, 月, 日)
        end_date: (年, 月, 日)
    """
    import copy
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = load_workbook(ledger_path)
    wb_cache = load_workbook(ledger_path, data_only=True)  # for reading cached values

    # 查找模板sheet
    template_sheet = None
    for s in wb.sheetnames:
        if '工资表' in s:
            template_sheet = wb[s]
            break

    if template_sheet is None:
        raise ValueError(f"未找到工资表模板sheet，现有sheet: {wb.sheetnames}")

    # 创建新sheet
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    new_sheet = wb.create_sheet(sheet_name)

    # 复制模板的前5行（标题行结构）
    project_title = template_sheet.cell(row=1, column=1).value or ''
    sy, sm, sd = start_date
    ey, em, ed = end_date
    date_str = f'{sy}年{sm:02d}月{sd:02d}日 - {ey}年{em:02d}月{ed:02d}日'

    # Row 1: 标题
    new_sheet.cell(row=1, column=1, value=project_title)
    new_sheet.merge_cells('A1:R1')
    c = new_sheet.cell(row=1, column=1)
    c.font = copy.copy(template_sheet.cell(row=1, column=1).font)
    c.alignment = copy.copy(template_sheet.cell(row=1, column=1).alignment)

    # Row 2: 空
    # Row 3: 项目名称 + 日期
    project_info = template_sheet.cell(row=3, column=1).value or ''
    new_sheet.cell(row=3, column=1, value=project_info)
    new_sheet.merge_cells('A3:Q3')
    c = new_sheet.cell(row=3, column=1)
    c.font = copy.copy(template_sheet.cell(row=3, column=1).font)
    c.alignment = copy.copy(template_sheet.cell(row=3, column=1).alignment)

    new_sheet.cell(row=3, column=18, value=date_str)
    c = new_sheet.cell(row=3, column=18)
    c.font = copy.copy(template_sheet.cell(row=3, column=18).font)
    c.alignment = copy.copy(template_sheet.cell(row=3, column=18).alignment)

    # Row 4: 分组标题（基础工资/加班/扣除项）
    for col in range(1, 19):
        src_cell = template_sheet.cell(row=4, column=col)
        dst_cell = new_sheet.cell(row=4, column=col)
        dst_cell.value = src_cell.value
        if src_cell.font:
            dst_cell.font = copy.copy(src_cell.font)
        if src_cell.alignment:
            dst_cell.alignment = copy.copy(src_cell.alignment)

    # Row 4 合并单元格
    for merge_range in template_sheet.merged_cells.ranges:
        if merge_range.min_row == 4:
            new_sheet.merge_cells(str(merge_range))

    # Row 5: 列标题
    col_headers = ['序号', '班组', '姓名', '工种', '出勤\n工日', '日工资',
                   '加班\n工时', '加班\n工资', '路费', '工资总额',
                   '预支费', '罚款', '其它', '项目部代付', '吴超付',
                   '未支付数', '领款人签字', '备注']
    for col_idx, header in enumerate(col_headers, 1):
        c = new_sheet.cell(row=5, column=col_idx, value=header)
        src_cell = template_sheet.cell(row=5, column=col_idx)
        if src_cell.font:
            c.font = copy.copy(src_cell.font)
        if src_cell.alignment:
            c.alignment = copy.copy(src_cell.alignment)
        if src_cell.border:
            c.border = copy.copy(src_cell.border)
        if src_cell.fill:
            c.fill = copy.copy(src_cell.fill)

    # 复制列宽
    for col_idx in range(1, 19):
        col_letter = get_column_letter(col_idx)
        if col_letter in template_sheet.column_dimensions:
            new_sheet.column_dimensions[col_letter].width = template_sheet.column_dimensions[col_letter].width

    # 填入数据（从 Row 6 开始）
    for row_idx, (_, row) in enumerate(salary_df.iterrows(), 6):
        new_sheet.cell(row=row_idx, column=1, value=row['序号'])
        new_sheet.cell(row=row_idx, column=2, value='')  # 班组留空
        new_sheet.cell(row=row_idx, column=3, value=row['姓名'])
        new_sheet.cell(row=row_idx, column=4, value=row['工种'])
        new_sheet.cell(row=row_idx, column=5, value=row['出勤工日'])
        new_sheet.cell(row=row_idx, column=6, value=row['日工资'] if row['日工资'] > 0 else None)
        new_sheet.cell(row=row_idx, column=7, value=row['加班工时'] if row['加班工时'] > 0 else None)
        new_sheet.cell(row=row_idx, column=8, value=row['加班工资'] if row['加班工资'] > 0 else None)
        # Col 9 路费留空
        new_sheet.cell(row=row_idx, column=10, value=row['工资总额'] if row['工资总额'] > 0 else None)
        # Col 11-14 预支费/罚款/其它/项目部代付 留空
        # Col 15 吴超付 留空
        new_sheet.cell(row=row_idx, column=16, value=row['未支付数'] if row['未支付数'] > 0 else None)
        # Col 17 领款人签字 留空
        new_sheet.cell(row=row_idx, column=18, value=row['备注'] if row.get('备注') else None)

        # 复制数据行样式（参考模板数据行）
        for col_idx in range(1, 19):
            c = new_sheet.cell(row=row_idx, column=col_idx)
            # 从模板第一个数据行复制样式
            src_cell = template_sheet.cell(row=6, column=col_idx)
            if src_cell.border:
                c.border = copy.copy(src_cell.border)
            if src_cell.alignment:
                c.alignment = copy.copy(src_cell.alignment)
            if src_cell.font:
                c.font = copy.copy(src_cell.font)

    # 合计行
    total_row = row_idx + 1
    new_sheet.cell(row=total_row, column=10,
                   value=round(salary_df['工资总额'].sum(), 1))
    new_sheet.cell(row=total_row, column=16,
                   value=round(salary_df['未支付数'].sum(), 1))

    for col_idx in range(1, 19):
        c = new_sheet.cell(row=total_row, column=col_idx)
        src_cell = template_sheet.cell(row=6, column=col_idx)
        if src_cell.border:
            c.border = copy.copy(src_cell.border)

    # 设置合计行字体加粗
    bold_font = Font(bold=True)
    new_sheet.cell(row=total_row, column=10).font = bold_font
    new_sheet.cell(row=total_row, column=16).font = bold_font

    # ==================== 更新年度工资汇总台账 ====================
    _update_annual_summary(wb, wb_cache, sheet_name, salary_df, start_date)

    wb.save(output_path)
    wb_cache.close()
    return output_path


def _update_annual_summary(wb, wb_cache, sheet_name: str, salary_df: pd.DataFrame, start_date: tuple):
    """
    将工资数据回写到"年度工资汇总台账"sheet

    规则: 按起始月份确定写入列 (3月11日-4月10日 → 3月工资列)
    列映射: N月工资 = Col(8 + 2*N), N月代付 = Col(9 + 2*N)

    写入实际数值（而非公式），确保手机端也能正确显示。

    Args:
        wb: 已打开的 openpyxl Workbook (for writing)
        wb_cache: 同文件以 data_only=True 打开的 Workbook (for reading cached values)
        sheet_name: 新增的工资表sheet名称
        salary_df: 工资汇总数据
        start_date: (年, 月, 日) 起始日期
    """
    import copy
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if '年度工资汇总台账' not in wb.sheetnames:
        return

    summary_sheet = wb['年度工资汇总台账']
    summary_cache = wb_cache['年度工资汇总台账']
    _, start_month, _ = start_date

    # 计算目标列: N月工资 = Col(8 + 2*N)
    wage_col = 8 + 2 * start_month  # e.g. 3月 → Col 14

    # 构建当前月工资查找: name → salary
    current_salary = {}
    for _, row in salary_df.iterrows():
        name = str(row['姓名']).strip()
        current_salary[name] = row['工资总额'] if row['工资总额'] > 0 else 0

    # 构建年度汇总中的姓名→行号映射 (Col C = column 3)
    existing_names = {}
    for row_idx in range(4, summary_sheet.max_row + 1):
        name_val = summary_sheet.cell(row=row_idx, column=3).value
        if name_val and str(name_val).strip() not in ('', '合计', '总计'):
            existing_names[str(name_val).strip()] = row_idx

    # 构建新工资表的姓名→行号映射 (用于公式引用)
    new_sheet = wb[sheet_name]
    new_sheet_names = {}
    for row_idx in range(6, new_sheet.max_row + 1):
        name_val = new_sheet.cell(row=row_idx, column=3).value
        if name_val:
            new_sheet_names[str(name_val).strip()] = row_idx

    # 找到合计行和总计行的位置
    total_row = None
    grand_total_row = None
    for row_idx in range(4, summary_sheet.max_row + 1):
        val = summary_sheet.cell(row=row_idx, column=1).value
        if val == '合计':
            total_row = row_idx
        elif val == '总计':
            grand_total_row = row_idx

    if not total_row:
        return

    # 找到合计行之前最后一个"小计"行（col3含"合"），用于确定合计的求和范围
    last_subtotal_row = 3  # 默认从数据起始行之前
    for r in range(4, total_row):
        c3 = summary_sheet.cell(row=r, column=3).value
        if c3 and '合' in str(c3).strip() and str(c3).strip() != '合计':
            last_subtotal_row = r

    # 用于复制新增行样式的参考行（合计行上一行）
    sample_data_row = total_row - 1

    def _get_person_total_salary(row_idx):
        """从缓存workbook读取某行的工资合计（各月工资之和）"""
        total = 0.0
        for m in range(1, 13):
            col = 8 + 2 * m
            if m == start_month:
                # 当前月用我们计算的数据
                name_val = summary_cache.cell(row=row_idx, column=3).value
                name = str(name_val).strip() if name_val else ''
                if name in current_salary:
                    total += current_salary[name]
                else:
                    v = summary_cache.cell(row=row_idx, column=col).value
                    if v and isinstance(v, (int, float)):
                        total += float(v)
            else:
                v = summary_cache.cell(row=row_idx, column=col).value
                if v and isinstance(v, (int, float)):
                    total += float(v)
        return round(total, 2)

    # 写入工资数据
    new_people = []  # 记录需要新增的人员
    for _, row in salary_df.iterrows():
        name = str(row['姓名']).strip()

        if name in existing_names:
            # 已有人员：写入实际数值到月份列
            target_row = existing_names[name]
            salary_val = row['工资总额'] if row['工资总额'] > 0 else 0
            summary_sheet.cell(row=target_row, column=wage_col, value=salary_val)
            # 更新工资合计为实际累加值
            summary_sheet.cell(row=target_row, column=5,
                               value=_get_person_total_salary(target_row))
        else:
            # 新增人员
            new_people.append((name, row))

    # 新增人员：在总计行之前插入（合计行保持不动）
    # 结构: 数据行 → 合计 → [新增人员] → 总计
    if new_people and grand_total_row:
        n = len(new_people)

        def _copy_row(src_row, dst_row, clear_src=False):
            """复制一行数据到另一行，处理合并单元格"""
            merged_ranges = list(summary_sheet.merged_cells.ranges)
            src_merged = [m for m in merged_ranges if m.min_row == src_row and m.max_row == src_row]
            # 先取消源行的合并
            for m in src_merged:
                summary_sheet.unmerge_cells(str(m))

            for col_idx in range(1, summary_sheet.max_column + 1):
                src_cell = summary_sheet.cell(row=src_row, column=col_idx)
                dst_cell = summary_sheet.cell(row=dst_row, column=col_idx)
                dst_cell.value = src_cell.value
                if src_cell.font:
                    dst_cell.font = copy.copy(src_cell.font)
                if src_cell.border:
                    dst_cell.border = copy.copy(src_cell.border)
                if src_cell.alignment:
                    dst_cell.alignment = copy.copy(src_cell.alignment)
                if clear_src:
                    src_cell.value = None

            # 在目标行重建合并
            for m in src_merged:
                new_range_str = f'{get_column_letter(m.min_col)}{dst_row}:{get_column_letter(m.max_col)}{dst_row}'
                summary_sheet.merge_cells(new_range_str)

        # 只移动总计行（合计行保持不动，新人员插在合计与总计之间）
        _copy_row(grand_total_row, grand_total_row + n, clear_src=True)

        # 填入新增人员（从 total_row+1 开始）
        for i, (name, row) in enumerate(new_people):
            new_row = total_row + 1 + i
            summary_sheet.cell(row=new_row, column=1, value=i + 1 + (total_row - 4))
            summary_sheet.cell(row=new_row, column=2, value='')
            summary_sheet.cell(row=new_row, column=3, value=name)
            summary_sheet.cell(row=new_row, column=4, value=row['工种'])
            # 代付合计 Col 6 — 新人员无代付数据，写 0
            summary_sheet.cell(row=new_row, column=6, value=0)
            # 未支付金额 Col 8 = 工资合计 - 代付 - 吴超付
            salary_val = row['工资总额'] if row['工资总额'] > 0 else 0
            summary_sheet.cell(row=new_row, column=5, value=salary_val)
            summary_sheet.cell(row=new_row, column=8, value=salary_val)

            # 写入当月工资
            summary_sheet.cell(row=new_row, column=wage_col, value=salary_val)

            # 复制样式
            for col_idx in range(1, summary_sheet.max_column + 1):
                src = summary_sheet.cell(row=sample_data_row, column=col_idx)
                dst = summary_sheet.cell(row=new_row, column=col_idx)
                if src.border:
                    dst.border = copy.copy(src.border)
                if src.alignment:
                    dst.alignment = copy.copy(src.alignment)
                if src.font:
                    dst.font = copy.copy(src.font)

        # 更新合计行：只求和最后一组数据（从小计行+1到合计行-1）
        group_start = last_subtotal_row + 1
        last_new_row = total_row + n
        for col_idx in range(5, 34):  # Col E ~ Col AG
            group_total = 0.0
            for r in range(group_start, total_row):
                v = summary_sheet.cell(row=r, column=col_idx).value
                if v and isinstance(v, (int, float)):
                    group_total += float(v)
            summary_sheet.cell(row=total_row, column=col_idx,
                               value=round(group_total, 2))

        # 更新总计行：原有值 + 新增人员总和
        if grand_total_row:
            new_grand_total_row = grand_total_row + n
            for col_idx in range(5, 34):
                # 读取原总计的缓存值
                cached = summary_cache.cell(row=grand_total_row, column=col_idx).value
                grand_val = float(cached) if cached and isinstance(cached, (int, float)) else 0.0
                # 加上新增人员的值
                for r in range(total_row + 1, last_new_row + 1):
                    v = summary_sheet.cell(row=r, column=col_idx).value
                    if v and isinstance(v, (int, float)):
                        grand_val += float(v)
                summary_sheet.cell(row=new_grand_total_row, column=col_idx,
                                   value=round(grand_val, 2))


# ============================================================
# 8. 上报表生成（适配新数据源）
# ============================================================

def generate_report_format(
    daily_df: pd.DataFrame,
    roster_path: str,
    output_path: str,
    attendance_paths: list = None,
    project_name: str = "新建西安至十堰高速铁路西安东站站房及相关工程XDZSG-2标段站房装修工程",
    company_name: str = "北京创新智成建筑劳务有限公司",
):
    """
    生成上报格式的Excel文件（花名册 + 考勤表）

    Args:
        daily_df: 每日考勤明细数据
        roster_path: 花名册文件路径
        output_path: 输出文件路径
        attendance_paths: 考勤文件路径列表（用于读取日期范围）
        project_name: 项目名称
        company_name: 分包单位名称
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    import os

    # 获取考勤日期范围
    date_range_start, date_range_end = None, None
    if attendance_paths:
        date_range_start, date_range_end = get_attendance_date_range(attendance_paths)

    if not date_range_start or not date_range_end:
        # 尝试从roster同目录查找考勤文件
        roster_dir = os.path.dirname(roster_path) if roster_path else '.'
        att_files = [os.path.join(roster_dir, f) for f in os.listdir(roster_dir)
                     if '员工刷卡记录' in f and (f.endswith('.xls') or f.endswith('.xlsx'))]
        if att_files:
            date_range_start, date_range_end = get_attendance_date_range(att_files)

    if date_range_start and date_range_end:
        sy, sm, sd = date_range_start
        ey, em, ed = date_range_end
        print(f"考勤日期范围: {sy}-{sm}-{sd} 到 {ey}-{em}-{ed}")
    else:
        # 从数据推断
        all_dates = sorted(daily_df['日期'].dropna().unique())
        sm, sd = 1, int(min(all_dates))
        em, ed = 1, int(max(all_dates))

    year = sy if date_range_start else 2026
    month = em if date_range_end else 1

    # 过滤有效员工
    valid_employees = daily_df[daily_df['当日工时'] > 0]['姓名'].unique()
    daily_df_filtered = daily_df[daily_df['姓名'].isin(valid_employees)].copy()

    # 读取花名册（新格式）
    roster_dict = parse_xdz_roster(roster_path)
    punch_employees = set(daily_df_filtered['姓名'].dropna().unique())

    # 构建花名册数据（仅出勤人员）
    roster_filtered = []
    for idx, name in enumerate(sorted(punch_employees), 1):
        info = roster_dict.get(name, {})
        roster_filtered.append({
            '编号': idx,
            '姓名': name,
            '性别': info.get('性别', ''),
            '工种': info.get('工种', ''),
            '身份证地址': info.get('身份证地址', ''),
            '身份证号码': info.get('身份证号码', ''),
            '电话号码': info.get('电话号码', ''),
            '合同编号': info.get('合同编号', ''),
        })

    # 计算日期列表
    if date_range_start and date_range_end and sm != em:
        # 跨月
        def get_month_days(y, m):
            if m == 2:
                return 29 if (y % 4 == 0 and y % 100 != 0) or (y % 400 == 0) else 28
            elif m in [4, 6, 9, 11]:
                return 30
            else:
                return 31

        month1_full = list(range(sd, get_month_days(sy, sm) + 1))
        month2_full = list(range(1, ed + 1))
        month_groups = [(f'{sm}月', month1_full), (f'{em}月', month2_full)]
        dates = month1_full + month2_full
    else:
        dates = list(range(sd, ed + 1))
        month_groups = [(f'{sm}月', dates)]

    # 创建工作簿
    wb = Workbook()
    wb.remove(wb.active)

    title_font = Font(bold=True, size=12)
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # ==================== Sheet 1: 花名册 ====================
    sheet1 = wb.create_sheet(f'花名册({month}月)')

    sheet1['A1'] = '劳务作业人员(含管理人员和作业人员)花名册'
    sheet1.merge_cells('A1:H1')
    sheet1['A1'].font = title_font
    sheet1['A1'].alignment = Alignment(horizontal='center', vertical='center')

    sheet1['A2'] = f'项目名称(全称):{project_name}'
    sheet1.merge_cells('A2:H2')
    sheet1['A2'].alignment = Alignment(horizontal='left', vertical='center')

    sheet1['A3'] = f'分包单位名称:{company_name}'
    sheet1.merge_cells('A3:E3')
    sheet1['F3'] = f'{year}年'
    sheet1['G3'] = f'{month}月'
    sheet1.merge_cells('F3:G3')

    headers1 = ['编号', '姓名', '性别', '工种(或岗位)', '家庭地址', '身份证号', '联系电话', '劳动合同编号']
    sheet1.append([])
    sheet1.append(headers1)

    for col_idx, header in enumerate(headers1, 1):
        cell = sheet1.cell(row=5, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    for r in roster_filtered:
        sheet1.append([
            r['编号'], r['姓名'], r.get('性别', ''), r['工种'],
            r.get('身份证地址', ''), r.get('身份证号码', ''),
            r.get('电话号码', ''), r.get('合同编号', '')
        ])

    for row in sheet1.iter_rows(min_row=6, max_row=sheet1.max_row):
        for cell in row:
            cell.border = thin_border

    col_widths1 = [6, 10, 6, 12, 30, 18, 12, 14]
    for idx, w in enumerate(col_widths1, 1):
        sheet1.column_dimensions[get_column_letter(idx)].width = w

    # ==================== Sheet 2: 考勤表 ====================
    sheet2 = wb.create_sheet(f'考勤表({year}.{month}月)')

    end_col = 3 + len(dates) + 3
    end_col_letter = get_column_letter(end_col)

    sheet2['A1'] = '劳务作业人员(含管理人员和作业人员)考勤表'
    sheet2.merge_cells(f'A1:{end_col_letter}1')
    sheet2['A1'].font = title_font
    sheet2['A1'].alignment = Alignment(horizontal='center', vertical='center')

    sheet2['A2'] = f'项目名称(全称):{project_name}'
    sheet2.merge_cells(f'A2:{end_col_letter}2')

    sheet2['A3'] = f'分包单位名称:{company_name}'
    sheet2.merge_cells(f'A3:{end_col_letter}3')

    # 表头行1: 月份
    header_row1 = ['编号', '姓名', '工种']
    for month_name, month_dates in month_groups:
        header_row1.append(month_name)
        header_row1.extend([''] * (len(month_dates) - 1))
    header_row1.extend(['', '', ''])
    sheet2.append(header_row1)

    # 合并月份单元格
    current_col = 4
    for month_name, month_dates in month_groups:
        if len(month_dates) > 1:
            month_end_col = current_col + len(month_dates) - 1
            sheet2.merge_cells(f'{get_column_letter(current_col)}4:{get_column_letter(month_end_col)}4')
        sheet2.cell(row=4, column=current_col).alignment = Alignment(horizontal='center', vertical='center')
        current_col += len(month_dates)

    # 表头行2: 日期
    header_row2 = ['', '', ''] + [str(d) for d in dates] + ['', '', '']
    sheet2.append(header_row2)

    # 表头行3: 加班/天数、合计、工人签字
    header_row3 = ['', '', ''] + [''] * len(dates) + ['加班/天数', '合计', '工人签字']
    sheet2.append(header_row3)

    for row_idx in [4, 5, 6]:
        for col_idx in range(1, sheet2.max_column + 1):
            cell = sheet2.cell(row=row_idx, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    # 数据行
    for emp in roster_filtered:
        base_row = sheet2.max_row + 1
        emp_name = emp['姓名']
        emp_job = emp['工种']

        emp_att = daily_df_filtered[daily_df_filtered['姓名'] == emp_name]

        date_hours = {}
        date_overtime = {}
        for _, att in emp_att.iterrows():
            if pd.notna(att['当日工时']) and att['当日工时'] > 0:
                date_hours[att['日期']] = att['当日工时']
                date_overtime[att['日期']] = att['加班工时']

        overtime_days = len([d for d, h in date_overtime.items() if h > 0])
        total_ot_hours = sum(date_overtime.values())
        合计 = overtime_days + round(total_ot_hours / 6, 1)

        # 出勤行
        row_data = [emp['编号'], emp_name, emp_job]
        for d in dates:
            row_data.append('√' if d in date_hours else '')
        row_data.extend([
            overtime_days if overtime_days > 0 else '',
            round(合计, 1) if 合计 > 0 else '',
            ''
        ])
        sheet2.append(row_data)

        # 加班行
        row_data2 = ['', '', '']
        for d in dates:
            row_data2.append(date_overtime[d] if d in date_overtime and date_overtime[d] > 0 else '')
        row_data2.extend([
            total_ot_hours if total_ot_hours > 0 else '',
            round(合计, 1) if 合计 > 0 else '',
            ''
        ])
        sheet2.append(row_data2)

        # 合并单元格
        end_col_idx = 3 + len(dates) + 3
        sheet2.merge_cells(f'A{base_row}:A{base_row+1}')
        sheet2.merge_cells(f'B{base_row}:B{base_row+1}')
        sheet2.merge_cells(f'C{base_row}:C{base_row+1}')
        sheet2.merge_cells(f'{get_column_letter(end_col_idx-1)}{base_row}:{get_column_letter(end_col_idx-1)}{base_row+1}')
        sheet2.merge_cells(f'{get_column_letter(end_col_idx)}{base_row}:{get_column_letter(end_col_idx)}{base_row+1}')

        sheet2[f'A{base_row}'].alignment = Alignment(horizontal='center', vertical='center')
        sheet2[f'B{base_row}'].alignment = Alignment(horizontal='center', vertical='center')
        sheet2[f'C{base_row}'].alignment = Alignment(horizontal='center', vertical='center')

    for row in sheet2.iter_rows(min_row=7, max_row=sheet2.max_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 列宽
    sheet2.column_dimensions['A'].width = 6
    sheet2.column_dimensions['B'].width = 10
    sheet2.column_dimensions['C'].width = 8
    for i, d in enumerate(dates):
        sheet2.column_dimensions[get_column_letter(4 + i)].width = 4
    ec = 3 + len(dates) + 3
    sheet2.column_dimensions[get_column_letter(ec-2)].width = 10
    sheet2.column_dimensions[get_column_letter(ec-1)].width = 10
    sheet2.column_dimensions[get_column_letter(ec)].width = 12

    wb.save(output_path)
    return output_path
